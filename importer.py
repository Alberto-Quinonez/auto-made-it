import csv
from copy import copy
from datetime import datetime
from typing import NamedTuple

import openpyxl

from config import BankConfig, BMOConfig, DESJConfig


class _BmoRow(NamedTuple):
    card: str
    tx_type: str
    date_str: str
    amount_str: str
    description: str


class _DESJRow(NamedTuple):
    account_type: str
    date: datetime
    code: float | None
    description: str
    withdrawal: float | None
    deposit: float | None
    interest: float | None
    capital: float | None


def import_csv_to_excel(csv_path: str, excel_path: str, bank: BankConfig) -> str:
    if isinstance(bank, BMOConfig):
        return _import_bmo(csv_path, excel_path, bank)
    elif isinstance(bank, DESJConfig):
        return _import_desj(csv_path, excel_path, bank)
    raise ValueError(f"Unknown bank config type: {type(bank)}")


# ── BMO ──────────────────────────────────────────────────────────────────────

def _import_bmo(csv_path: str, excel_path: str, bank: BMOConfig) -> str:
    csv_rows = _parse_bmo_csv(csv_path, bank)

    wb_ro = openpyxl.load_workbook(excel_path, data_only=True)
    ws_ro = wb_ro[bank.excel_sheet]
    insert_row = _find_insert_row(ws_ro, bank.excel_date, bank.excel_data_start_row)
    existing = _bmo_duplicate_set(ws_ro, insert_row, bank)
    last_balance = _get_last_balance(ws_ro, insert_row - 1, bank.excel_data_start_row,
                                     bank.excel_balance, bank.excel_dr, bank.excel_cr)
    wb_ro.close()

    new_rows, skipped = [], 0
    for row in csv_rows:
        try:
            key = (int(row.date_str), float(row.amount_str), row.description)
        except ValueError:
            continue
        if key in existing:
            skipped += 1
        else:
            new_rows.append(row)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb[bank.excel_sheet]

    if new_rows:
        insert_row = _find_insert_row(ws, bank.excel_date, bank.excel_data_start_row)
        ws.insert_rows(insert_row, amount=len(new_rows))
        template_row = insert_row - 1
        max_col = ws.max_column

        for i, row in enumerate(new_rows):
            r = insert_row + i
            amt = float(row.amount_str)
            last_balance = round(last_balance + amt, 2)
            _copy_row_style(ws, template_row, r, max_col)
            ws.cell(row=r, column=bank.excel_je).value = None
            ws.cell(row=r, column=bank.excel_card).value = row.card
            ws.cell(row=r, column=bank.excel_type).value = row.tx_type
            ws.cell(row=r, column=bank.excel_date).value = int(row.date_str)
            ws.cell(row=r, column=bank.excel_amount).value = amt
            ws.cell(row=r, column=bank.excel_desc).value = row.description
            ws.cell(row=r, column=bank.excel_dr).value = abs(amt) if row.tx_type == "CREDIT" else None
            ws.cell(row=r, column=bank.excel_cr).value = amt if row.tx_type == "DEBIT" else None
            ws.cell(row=r, column=bank.excel_balance).value = last_balance

    wb.save(excel_path)
    return f"Done — {len(new_rows)} row(s) added, {skipped} skipped."


def _parse_bmo_csv(csv_path: str, bank: BMOConfig) -> list[_BmoRow]:
    rows = []
    header_found = False
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row in reader:
            if not header_found:
                if len(row) > bank.header_col_index and row[bank.header_col_index].strip() == bank.header_col_value:
                    header_found = True
                continue
            if len(row) < 5 or not row[bank.csv_date].strip():
                continue
            rows.append(_BmoRow(
                card=row[bank.csv_card].strip(),
                tx_type=row[bank.csv_type].strip(),
                date_str=row[bank.csv_date].strip(),
                amount_str=row[bank.csv_amount].strip(),
                description=row[bank.csv_desc].strip(),
            ))
    return rows


def _bmo_duplicate_set(ws, insert_row: int, bank: BMOConfig) -> set:
    seen = set()
    for row in range(bank.excel_data_start_row, insert_row):
        date = ws.cell(row=row, column=bank.excel_date).value
        amt = ws.cell(row=row, column=bank.excel_amount).value
        desc = ws.cell(row=row, column=bank.excel_desc).value
        if date is not None and amt is not None and desc:
            try:
                seen.add((int(date), float(amt), str(desc).strip()))
            except (ValueError, TypeError):
                pass
    return seen


# ── DESJ ─────────────────────────────────────────────────────────────────────

def _import_desj(csv_path: str, excel_path: str, bank: DESJConfig) -> str:
    csv_rows = _parse_desj_csv(csv_path, bank)

    wb_ro = openpyxl.load_workbook(excel_path, data_only=True)
    ws_ro = wb_ro[bank.excel_sheet]
    insert_row = _find_insert_row(ws_ro, bank.excel_date, bank.excel_data_start_row)
    existing = _desj_duplicate_set(ws_ro, insert_row, bank)
    last_cash = _get_balance_simple(ws_ro, insert_row - 1, bank.excel_data_start_row, bank.excel_balance)
    last_mortgage = _get_balance_simple(ws_ro, insert_row - 1, bank.excel_data_start_row, bank.excel_mortgage_balance)
    wb_ro.close()

    new_rows, skipped = [], 0
    for row in csv_rows:
        key = (row.date, row.description)
        if key in existing:
            skipped += 1
        else:
            new_rows.append(row)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb[bank.excel_sheet]

    if new_rows:
        insert_row = _find_insert_row(ws, bank.excel_date, bank.excel_data_start_row)

        grey_fill = _find_grey_fill(ws, bank.excel_data_start_row, insert_row, bank.excel_balance)
        blank_fill = copy(ws.cell(row=bank.excel_data_start_row, column=bank.excel_date).fill)

        ws.insert_rows(insert_row, amount=len(new_rows))
        template_row = insert_row - 1
        max_col = ws.max_column

        for i, row in enumerate(new_rows):
            r = insert_row + i
            _copy_row_style(ws, template_row, r, max_col)
            ws.cell(row=r, column=bank.excel_je).value = None

            bal_cell = ws.cell(row=r, column=bank.excel_balance)
            if row.account_type == "LN1" and grey_fill:
                bal_cell.fill = grey_fill
            else:
                bal_cell.fill = blank_fill

            ws.cell(row=r, column=bank.excel_date).value = row.date
            ws.cell(row=r, column=bank.excel_code).value = row.code
            ws.cell(row=r, column=bank.excel_desc).value = row.description

            if row.account_type == "PCA":
                if row.withdrawal:
                    ws.cell(row=r, column=bank.excel_withdrawal).value = row.withdrawal
                    last_cash = round(last_cash - row.withdrawal, 2)
                if row.deposit:
                    ws.cell(row=r, column=bank.excel_deposit).value = row.deposit
                    last_cash = round(last_cash + row.deposit, 2)
                ws.cell(row=r, column=bank.excel_balance).value = last_cash

            elif row.account_type == "LN1":
                if row.interest:
                    ws.cell(row=r, column=bank.excel_interest).value = row.interest
                if row.capital:
                    ws.cell(row=r, column=bank.excel_capital).value = row.capital
                    last_mortgage = round(last_mortgage - row.capital, 2)
                ws.cell(row=r, column=bank.excel_mortgage_balance).value = last_mortgage

    wb.save(excel_path)
    return f"Done — {len(new_rows)} row(s) added, {skipped} skipped."


def _parse_desj_csv(csv_path: str, bank: DESJConfig) -> list[_DESJRow]:
    expected_cols = bank.csv_capital + 2
    rows = []
    with open(csv_path, newline="", encoding="latin-1") as f:
        reader = csv.reader(f)
        for raw in reader:
            n = len(raw)
            if n < expected_cols:
                continue
            if n > expected_cols:
                # Description field contains unquoted commas — reconstruct it
                extra = n - expected_cols
                desc = ",".join(raw[bank.csv_desc : bank.csv_desc + 1 + extra])
                row = raw[: bank.csv_desc] + [desc] + raw[bank.csv_desc + 1 + extra :]
            else:
                row = raw
            if not row[bank.csv_date].strip():
                continue
            try:
                date = datetime.strptime(row[bank.csv_date].strip(), "%Y/%m/%d")
            except ValueError:
                continue
            rows.append(_DESJRow(
                account_type=row[bank.csv_account_type].strip(),
                date=date,
                code=_to_float(row[bank.csv_code]),
                description=row[bank.csv_desc].strip(),
                withdrawal=_to_float(row[bank.csv_withdrawal]),
                deposit=_to_float(row[bank.csv_deposit]),
                interest=_to_float(row[bank.csv_interest]),
                capital=_to_float(row[bank.csv_capital]) if len(row) > bank.csv_capital else None,
            ))
    return rows


def _desj_duplicate_set(ws, insert_row: int, bank: DESJConfig) -> set:
    seen = set()
    for row in range(bank.excel_data_start_row, insert_row):
        date = ws.cell(row=row, column=bank.excel_date).value
        desc = ws.cell(row=row, column=bank.excel_desc).value
        if date is not None and desc:
            seen.add((date, str(desc).strip()))
    return seen


# ── Shared helpers ────────────────────────────────────────────────────────────

def _find_insert_row(ws, date_col: int, data_start_row: int) -> int:
    end = max(ws.max_row, data_start_row) + 2
    for row in range(data_start_row, end):
        if ws.cell(row=row, column=date_col).value is None:
            return row


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format


def _scan_for_balance(ws, from_row: int, to_row: int, balance_col: int) -> float | None:
    for row in range(from_row, to_row - 1, -1):
        val = ws.cell(row=row, column=balance_col).value
        if val is not None:
            try:
                return float(val)
            except (ValueError, TypeError):
                continue
    return None


def _get_balance_simple(ws, last_row: int, data_start_row: int, balance_col: int) -> float:
    return _scan_for_balance(ws, last_row, data_start_row, balance_col) or 0.0


def _get_last_balance(ws, last_row: int, data_start_row: int,
                      balance_col: int, dr_col: int, cr_col: int) -> float:
    cached = _scan_for_balance(ws, last_row, data_start_row, balance_col)
    if cached is not None:
        return cached

    # Formulas not cached — compute from opening balance + all DR/CR
    opening = _scan_for_balance(ws, data_start_row - 1, 1, balance_col) or 0.0
    balance = opening
    for row in range(data_start_row, last_row + 1):
        for col in (dr_col, cr_col):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                try:
                    balance += float(val)
                except (ValueError, TypeError):
                    pass
    return round(balance, 2)


def _find_grey_fill(ws, data_start_row: int, insert_row: int, balance_col: int):
    for row in range(data_start_row, insert_row):
        cell = ws.cell(row=row, column=balance_col)
        if cell.fill.fill_type == "solid":
            return copy(cell.fill)
    return None


def _to_float(val: str) -> float | None:
    if not val or not val.strip():
        return None
    try:
        return float(val.strip().replace(",", ""))
    except ValueError:
        return None
